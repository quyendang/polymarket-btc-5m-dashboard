"""Backtest execution and XLSX rendering for the dashboard worker."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

import backtest
import compare_runs


def execute_backtest(hours: int, starting: float, min_bet: float) -> dict:
    candles = backtest.fetch_candles(hours=hours)
    evals = compare_runs.build_windows(candles)
    runs = [
        compare_runs.simulate(evals, mode, threshold, starting, min_bet)
        for mode in compare_runs.MODES
        for threshold in compare_runs.CONF_THRESHOLDS
    ]
    best = max(runs, key=lambda item: item.roi) if runs else None
    configs = [
        {
            "mode": item.mode,
            "threshold": item.threshold,
            "trades": item.trades,
            "wins": item.wins,
            "win_rate": item.wins / item.trades if item.trades else 0.0,
            "final_bankroll": item.final_bankroll,
            "roi": item.roi,
            "max_drawdown": item.max_drawdown,
            "curve": item.curve,
        }
        for item in runs
    ]
    return {
        "hours": hours,
        "windows_count": len(evals),
        "starting_bankroll": starting,
        "min_bet": min_bet,
        "configs": configs,
        "best": None if best is None else {
            "mode": best.mode,
            "threshold": best.threshold,
            "trades": best.trades,
            "wins": best.wins,
            "win_rate": best.wins / best.trades if best.trades else 0.0,
            "final_bankroll": best.final_bankroll,
            "roi": best.roi,
            "max_drawdown": best.max_drawdown,
            "trade_log": best.trade_log,
        },
    }


def workbook_bytes(payload: dict) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    headers = ["Mode", "Conf Threshold", "Trades", "Wins", "Win Rate",
               "Final Bankroll", "ROI", "Max Drawdown"]
    summary.append(headers)
    for cell in summary[1]:
        cell.font = Font(bold=True)
    for item in sorted(payload.get("configs", []), key=lambda row: row["roi"], reverse=True):
        summary.append([
            item["mode"], item["threshold"], item["trades"], item["wins"],
            item["win_rate"], item["final_bankroll"], item["roi"], item["max_drawdown"],
        ])
    summary.append([])
    summary.append([
        f"Windows evaluated: {payload.get('windows_count', 0)}",
        f"Lookback: {payload.get('hours', 0)}h",
        "Signal @ T-60s",
    ])

    trades = workbook.create_sheet("Best Config Trades")
    best = payload.get("best") or {}
    trade_log = best.get("trade_log") or []
    if trade_log:
        trades.append([
            f"Best: {best['mode']} @ {best['threshold']} · ROI {best['roi']:.2%}"
        ])
        columns = list(trade_log[0].keys())
        trades.append(columns)
        for cell in trades[2]:
            cell.font = Font(bold=True)
        for row in trade_log:
            trades.append([row.get(column) for column in columns])
    else:
        trades.append(["No trades"])

    curves = workbook.create_sheet("Bankroll Curves")
    configs = payload.get("configs", [])
    curves.append(["Trade #"] + [f"{item['mode']}@{item['threshold']}" for item in configs])
    for cell in curves[1]:
        cell.font = Font(bold=True)
    max_length = max((len(item.get("curve", [])) for item in configs), default=0)
    for index in range(max_length):
        curves.append([index + 1] + [
            item["curve"][index] if index < len(item.get("curve", [])) else None
            for item in configs
        ])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
