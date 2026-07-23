"""Backtest the strategy across many configs and write an Excel comparison.

For every historical 5-minute window we run the real strategy.analyze() and
simulate a trade under the delta-based pricing model, then sweep:

    9 confidence thresholds  x  3 modes (flat, safe, aggressive)  = 27 configs

Output workbook (3 sheets):
    Summary          — one row per config: trades, win rate, ROI, max drawdown
    Best Config Trades — trade-by-trade log of the highest-ROI config
    Bankroll Curves  — bankroll after each trade, one column per config

    python compare_runs.py --hours 72 --output results.xlsx

HONESTY NOTE — snapshot timing:
The live bot fires at T-10s. With 1-minute candles the finest pre-close
snapshot is T-60s (the close of the window's 4th minute); using the 5th
minute's close as "current price" would equal the resolution price and make
direction prediction trivially perfect. So the backtest evaluates the signal at
T-60s and resolves on the 5th minute's close. Real T-10s behavior sits between
this and perfect; treat backtest win-rates as a floor, not a promise.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass

import backtest
import bot
import pricing
import strategy

CONF_THRESHOLDS = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]
MODES = ["flat", "safe", "aggressive"]
CANDLE_LOOKBACK = 30
WINDOW = 300


@dataclass
class WindowEval:
    window_ts: int
    signal: strategy.Signal
    delta_pct: float
    direction: str
    actual_up: bool


def build_windows(candles: list) -> list[WindowEval]:
    """Group 1m candles into complete 5-min windows and evaluate each.

    Signal is computed at T-60s (4th candle close); outcome from the 5th
    candle close vs the window open.
    """
    by_time = {c.open_time: c for c in candles}
    chrono = [by_time[t] for t in sorted(by_time)]
    index = {c.open_time: i for i, c in enumerate(chrono)}

    evals: list[WindowEval] = []
    for c in chrono:
        w = c.open_time - (c.open_time % WINDOW)
        if c.open_time != w:
            continue  # only start at window boundaries
        needed = [w, w + 60, w + 120, w + 180, w + 240]
        if not all(t in by_time for t in needed):
            continue
        window_open = by_time[w].open
        eval_candle = by_time[w + 180]      # closes at w+240 == T-60s
        outcome_candle = by_time[w + 240]   # closes at w+300 == T-0

        i = index[w + 180]
        slice_ = chrono[max(0, i - CANDLE_LOOKBACK + 1): i + 1]
        current = eval_candle.close
        sig = strategy.analyze(slice_, window_open, current, ticks=None)
        delta_pct = (current - window_open) / window_open * 100 if window_open else 0.0
        direction = bot.resolve_direction(sig, delta_pct)
        actual_up = outcome_candle.close >= window_open
        evals.append(WindowEval(w, sig, delta_pct, direction, actual_up))
    return evals


@dataclass
class RunResult:
    mode: str
    threshold: float
    trades: int
    wins: int
    final_bankroll: float
    roi: float
    max_drawdown: float
    curve: list          # bankroll after each trade
    trade_log: list      # dicts, for the best-config sheet


def simulate(evals: list[WindowEval], mode: str, threshold: float,
             starting: float, min_bet: float) -> RunResult:
    bankroll = starting
    has_traded = False
    trades = wins = 0
    peak = starting
    max_dd = 0.0
    curve = []
    log = []

    for ev in evals:
        if ev.signal.confidence < threshold:
            continue                      # threshold skip (backtest only)
        if bankroll < min_bet:
            break                         # busted

        if mode == "flat":
            bet = min(min_bet, bankroll)
        else:
            bet = bot.bet_size(mode, bankroll, starting, min_bet, has_traded)
        if bet <= 0:
            break

        entry, shares, won, pnl = bot.score_dry_run(
            ev.direction, bet, ev.delta_pct, ev.actual_up)
        bankroll += pnl
        has_traded = True
        trades += 1
        wins += 1 if won else 0

        peak = max(peak, bankroll)
        if peak > 0:
            max_dd = max(max_dd, (peak - bankroll) / peak)
        curve.append(round(bankroll, 4))
        log.append({
            "window_ts": ev.window_ts,
            "direction": ev.direction,
            "score": ev.signal.score,
            "confidence": round(ev.signal.confidence, 3),
            "delta_pct": round(ev.delta_pct, 4),
            "entry": entry,
            "bet": round(bet, 4),
            "shares": round(shares, 3),
            "won": won,
            "pnl": pnl,
            "bankroll": round(bankroll, 4),
        })

    roi = (bankroll - starting) / starting if starting else 0.0
    return RunResult(mode, threshold, trades, wins, round(bankroll, 4),
                     round(roi, 4), round(max_dd, 4), curve, log)


def write_excel(results: list[RunResult], evals_count: int, hours: int,
                output: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    headers = ["Mode", "Conf Threshold", "Trades", "Wins", "Win Rate",
               "Final Bankroll", "ROI", "Max Drawdown"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in sorted(results, key=lambda x: x.roi, reverse=True):
        wr = (r.wins / r.trades) if r.trades else 0.0
        ws.append([r.mode, r.threshold, r.trades, r.wins, round(wr, 4),
                   r.final_bankroll, r.roi, r.max_drawdown])
    ws.append([])
    ws.append([f"Windows evaluated: {evals_count}", f"Lookback: {hours}h",
               "Signal @ T-60s (see file docstring)"])

    # Sheet 2: Best Config Trades
    best = max(results, key=lambda x: x.roi) if results else None
    ws2 = wb.create_sheet("Best Config Trades")
    if best and best.trade_log:
        ws2.append([f"Best config: mode={best.mode} threshold={best.threshold} "
                    f"ROI={best.roi:.2%} ({best.wins}/{best.trades})"])
        cols = list(best.trade_log[0].keys())
        ws2.append(cols)
        for cell in ws2[2]:
            cell.font = Font(bold=True)
        for row in best.trade_log:
            ws2.append([row[c] for c in cols])
    else:
        ws2.append(["No trades in best config."])

    # Sheet 3: Bankroll Curves (one column per config)
    ws3 = wb.create_sheet("Bankroll Curves")
    labels = [f"{r.mode}@{r.threshold}" for r in results]
    ws3.append(["Trade #"] + labels)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    max_len = max((len(r.curve) for r in results), default=0)
    for i in range(max_len):
        row = [i + 1]
        for r in results:
            row.append(r.curve[i] if i < len(r.curve) else None)
        ws3.append(row)

    wb.save(output)


def main() -> None:
    p = argparse.ArgumentParser(description="Backtest strategy across configs")
    p.add_argument("--hours", type=int, default=72)
    p.add_argument("--output", default="results.xlsx")
    p.add_argument("--starting", type=float, default=100.0)
    p.add_argument("--min-bet", type=float, default=1.0)
    args = p.parse_args()

    print(f"Fetching {args.hours}h of 1m candles...")
    candles = backtest.fetch_candles(hours=args.hours)
    print(f"  {len(candles)} candles")
    evals = build_windows(candles)
    print(f"  {len(evals)} complete 5-min windows evaluated")

    results = []
    for mode in MODES:
        for thr in CONF_THRESHOLDS:
            results.append(simulate(evals, mode, thr, args.starting, args.min_bet))

    write_excel(results, len(evals), args.hours, args.output)

    best = max(results, key=lambda x: x.roi)
    print(f"\nBest: mode={best.mode} threshold={best.threshold} "
          f"ROI={best.roi:.2%} final=${best.final_bankroll:.2f} "
          f"({best.wins}/{best.trades} = "
          f"{best.wins/best.trades:.0%})" if best.trades else "no trades")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
